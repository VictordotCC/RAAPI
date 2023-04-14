"""Helpers Methods"""
import datetime
import requests

from fastkml import kml
import openpyxl as xl

import config


def get_weather_info(proyecto):
    """get wind speed and direction from openweather api"""

    url = (config.OW_URL + "lat=" + str(proyecto["Latitud"]) + "&lon="
        + str(proyecto["Longitud"]) + "&appid=" + config.OW_APIKEY)
    response = requests.get(url, timeout=30)
    data = response.json()
    result = {}
    for i in range(len(data["list"])):
        utc_time=datetime.datetime.fromtimestamp(data["list"][i]["dt"])
        chile_time=((utc_time+datetime.timedelta(hours=config.CHILETIMEZONE))
                    .strftime("%d-%m-%Y %H:%M:%S"))
        wind_speed = data["list"][i]["wind"]["speed"]
        wind_direction = data["list"][i]["wind"]["deg"]
        result.update({chile_time: {"wind_speed": wind_speed, "wind_direction": wind_direction}})
    return result

def leer_kml(kml_file):
    """Read a kml file and return the coordinates"""

    #check if kml_file is a file or a string
    if isinstance(kml_file, str):
        kml_file = kml_file.encode('utf-8')
    else:
        kml_file = kml_file.read()

    with open(kml_file, 'rb') as f:
        doc = f.read()
        k = kml.KML()
        k.from_string(doc)
        k.from_string(doc)
   
        features = list(k.features())
        puntos = []
        for feature in features:
            for group in feature.features():
                for placemark in group.features():
                    coordenadas = [placemark.geometry.x, placemark.geometry.y]
                    if coordenadas not in puntos:
                        puntos.append(coordenadas)
    return puntos

def get_time():
    """Get the current time"""
    return datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S")

def search_in_xlsx(id_proyecto, n_ag, n_receptor):
    """Search in the xlsx file the values of the receptor and aero generator"""
    file_path = "meds/" + str(id_proyecto) + ".xlsx"
    wb = xl.load_workbook(file_path)
    par_mediciones = []
    for sheet in wb.worksheets:
        title = sheet.title.split("-")
        vel_viento = title[0]
        #indexing receptores
        for cell in sheet.iter_cols(min_row=1, max_row=1):
            if cell[0].value == n_receptor:
                receptor_col = cell[0].column
                break
            else:
                receptor_col = None
        if receptor_col is None:
            break
        #search the n_ag in the sheet
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == n_ag:
                    row_num = cell.row
                    par_mediciones.append({"vel_viento": title[0],
                                            "angulo": sheet.cell(row=row_num, column=1).value,
                                            "valor": (str(sheet.cell(row=row_num, column=receptor_col).value)).replace(",", ".")})
                    break
    if len(par_mediciones) == 0:
        return None
    #print(par_mediciones)
    return par_mediciones



        




#search_in_xlsx("642851f426b13c3e932632f9", "AG1", "R1")

    



#x = {"Latitud": -33.4569, "Longitud": -70.6483}
#r = get_weather_info(x)
